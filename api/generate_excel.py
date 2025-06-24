import os
import io
import json
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

def handler(request):
    if request.method != "POST":
        return (400, {"Content-Type": "text/plain"}, b"Only POST allowed")
    
    try:
        body = request.body.decode() if isinstance(request.body, bytes) else request.body
        data = json.loads(body)
        products = data.get("products", [])
    except Exception as e:
        return (400, {"Content-Type": "text/plain"}, f"Error parsing body: {str(e)}".encode())
    
    # Ruta absoluta al template (¡ojo aquí!)
    template_path = os.path.join(os.path.dirname(__file__), "../templates/peddo_template.xlsx")
    template_path = os.path.abspath(template_path)
    if not os.path.exists(template_path):
        return (500, {"Content-Type": "text/plain"}, f"Template not found: {template_path}".encode())
    
    wb = load_workbook(template_path)
    ws = wb['PEDIDO']  # El nombre debe coincidir

    # Limpiá la hoja antes de agregar (importante para serverless)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.value = None

    # INSERTAR DATOS desde la fila 2
    for idx, prod in enumerate(products, start=2):
        ws[f"A{idx}"] = 1  # Cliente ID
        ws[f"B{idx}"] = prod.get('product_id')
        ws[f"C{idx}"] = ""  # Cantidad vacía
        ws[f"D{idx}"] = round(prod.get('price', 0) * 0.4, 2)  # Precio con 60% descuento
        ws[f"E{idx}"] = prod.get('display_name', '')  # Producto limpio
        ws[f"F{idx}"] = prod.get('qty_available', 0)  # Stock

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    headers = {
        "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "Content-Disposition": f"attachment; filename=pedido_stock.xlsx"
    }
    return (200, headers, output.getvalue())
